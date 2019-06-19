import xlrd
import xlwt
import copy
import xlutils
from openpyxl import load_workbook
import csv
import pandas as pd
import sys
import os
#ead csv
from os import listdir


a=1
try:
    if a == '1':
        print("test")
    else

except:
    print("t")

'''
dirs = os.listdir( '/Users/ian/Downloads/TEST' )

# This would print all the files and directories
for file in dirs:
   print (file)
'''

'''
#----------------------------------------------------------------------------------------------
#所有徑向、軸向數據
for i in range(1, 6):
    filepath_in = app.config['UPLOAD_FOLDER'] + "/A" + i + "0.csv"
    filepath_out = '/Users/ian/Downloads/test_data2' + "/A" + str(i) + "0.xlsx"
    pd.read_csv(filepath_in, delimiter=",").to_excel(filepath_out, index=False)
for i in range(1, 7):
    filepath_in = '/Users/ian/Downloads/test_data2' + "/B" + str(i) + "0.csv"
    filepath_out = '/Users/ian/Downloads/test_data2' + "/B" + str(i) + "0.xlsx"
    pd.read_csv(filepath_in, delimiter=",").to_excel(filepath_out, index=False)
cnt=4
filename = '/Users/ian/Downloads/test_data2/test2.xlsx'
wb = load_workbook(filename)
ws = wb.active
for i in range(1, 6):
    #Load excel data
    new_excel = '/Users/ian/Downloads/test_data2' + "/A" + str(i) + "0.xlsx"
    wb2 = load_workbook(new_excel)
    ws2 = wb2.active
    #轉速
    speed = ws2.cell(row=57, column=2).value[0:3]
    #徑向,軸向
    X_value = ws2.cell(row=59, column=2).value[0:4]
    Y_value = ws2.cell(row=59, column=3).value[0:4]
    cell_value = ws.cell(row=6, column=cnt).value
    ws.cell(row=6, column=cnt, value=str(cell_value)+'('+speed+')')
    ws.cell(row=11, column=cnt, value=X_value)
    ws.cell(row=12, column=cnt, value=Y_value)
    cnt = cnt+1
    wb.save("/Users/ian/Downloads/test_data2/report.xlsx")
cnt=10
for i in range(1, 7):
    #Load excel data
    new_excel = '/Users/ian/Downloads/test_data2' + "/B" + str(i) + "0.xlsx"
    wb2 = load_workbook(new_excel)
    ws2 = wb2.active
    #轉速
    speed = ws2.cell(row=57, column=2).value[0:3]
    #徑向,軸向
    X_value = ws2.cell(row=59, column=2).value[0:4]
    Y_value = ws2.cell(row=59, column=3).value[0:4]
    cell_value = ws.cell(row=6, column=cnt).value
    ws.cell(row=6, column=cnt, value=str(cell_value)+'('+speed+')')
    ws.cell(row=11, column=cnt, value=X_value)
    ws.cell(row=12, column=cnt, value=Y_value)
    cnt = cnt+1
    wb.save("/Users/ian/Downloads/test_data2/excel_test.xlsx")
#----------------------------------------------------------------------------------------------
'''
'''
with open(csvname, newline='',errors='ignore') as f:
    rows = csv.reader(f)
    for row in rows:
        print(row)
        #取出溫度
        Temp1 = row[1]
        Temp2 = row[3]
        InsertTemp = Temp1[0:4]+','+Temp2[0:4]
'''
#SUCESS Insert Value to Excel
'''
Todo:儲存格的大小調整

filename = '/Users/ian/Downloads/test.xlsx'
wb = load_workbook(filename)
#print(wb.sheetnames)
LoudFileName = '/Users/ian/Downloads/LA01test.csv'
ws = wb.active

EleFileName = '/Users/ian/Downloads/EA01.csv'
with open(EleFileName, newline='',errors='ignore') as e:
    reader = csv.reader(x.replace('\0', '') for x in e)
    for ro in reader:
        print(ro)

'''
#塞入數值
'''
ws=wb.active

ws.cell(row=7,column=4,value=InsertTemp)
print(ws['D9'].value)
wb.save('/Users/ian/Downloads/Openpyxl_test.xlsx')
'''
'''
    for line in f:
        print(line.decode(errors='ignore'))


def loud():
    #取出噪音
    with open(LoudFileName, newline='', errors='ignore') as l:
        loudrows = csv.reader(l)
        count=4
        for loudrow in loudrows:
            print(loudrow)
            LoudTemp = loudrow[1]
            print(LoudTemp)
            ws.cell(row=10,column=count,value=float(LoudTemp))
            count = count+1
    wb.save('/Users/ian/Downloads/Openpyxl_test.xlsx')
'''