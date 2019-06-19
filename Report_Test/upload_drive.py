import os, csv
from app import app
from flask import Flask, flash, request, redirect, render_template, send_file, send_from_directory
from werkzeug.utils import secure_filename
from openpyxl import load_workbook
import pandas as pd
from shutil import copyfile
import flask_excel as excel

ALLOWED_EXTENSIONS = set(['csv'])
#CSV to EXCEL
def convert():
    for i in range(1, 6):

        try:
            filepath_in = app.config['UPLOAD_FOLDER']+"/A"+str(i)+"0.csv"
            filepath_out = app.config['UPLOAD_FOLDER']+"/A"+str(i)+"0.xlsx"
            pd.read_csv(filepath_in, delimiter=",").to_excel(filepath_out, index=False)
        except :
            pass

        for i in range(1, 7):
            try:
                filepath_in = app.config['UPLOAD_FOLDER']+"/B"+str(i)+"0.csv"
                filepath_out = app.config['UPLOAD_FOLDER']+"/B"+str(i)+"0.xlsx"
                pd.read_csv(filepath_in, delimiter=",").to_excel(filepath_out, index=False)
            except :
                pass
#震動
def InsertA():
    cnt = 4
    filename = app.config['UPLOAD_FOLDER']+'/test2.xlsx'
    wb = load_workbook(filename)
    ws = wb.active
    for i in range(1, 6):
        try:
        # Load excel data
            new_excel = app.config['UPLOAD_FOLDER'] + "/A" + str(i) + "0.xlsx"
            wb2 = load_workbook(new_excel)
            ws2 = wb2.active
            # 轉速
            speed = ws2.cell(row=57, column=2).value[0:3]
            # 徑向,軸向
            X_value = ws2.cell(row=59, column=2).value[0:4]
            Y_value = ws2.cell(row=59, column=3).value[0:4]
            cell_value = ws.cell(row=6, column=cnt).value
            ws.cell(row=6, column=cnt, value=str(cell_value) + '(' + speed + ')')
            ws.cell(row=11, column=cnt, value=X_value)
            ws.cell(row=12, column=cnt, value=Y_value)
            cnt = cnt + 1
        except:
            pass
        wb.save(app.config['UPLOAD_FOLDER'] + '/report.xlsx')
    cnt = 10
    for i in range(1, 7):
        try:
            # Load excel data
            new_excel = app.config['UPLOAD_FOLDER'] + "/B" + str(i) + "0.xlsx"
            wb2 = load_workbook(new_excel)
            ws2 = wb2.active
            # 轉速
            speed = ws2.cell(row=57, column=2).value[0:3]
            # 徑向,軸向
            X_value = ws2.cell(row=59, column=2).value[0:4]
            Y_value = ws2.cell(row=59, column=3).value[0:4]
            cell_value = ws.cell(row=6, column=cnt).value
            ws.cell(row=6, column=cnt, value=str(cell_value) + '(' + speed + ')')
            ws.cell(row=11, column=cnt, value=X_value)
            ws.cell(row=12, column=cnt, value=Y_value)
            cnt = cnt + 1
        except:
            pass
        wb.save(app.config['UPLOAD_FOLDER'] + '/report.xlsx')

#取出噪音
def loud():
    filename = app.config['UPLOAD_FOLDER']+'/report.xlsx'
    wb = load_workbook(filename)
    ws = wb.active
    LoudFileNameA = app.config['UPLOAD_FOLDER']+'/LA1.csv'
    try:
        with open(LoudFileNameA, newline='', errors='ignore') as la:
            loudrows = csv.reader(la)
            count=4
            for loudrow in loudrows:
                try:
                    LoudTemp = loudrow[1]
                    ws.cell(row=10, column=count, value=float(LoudTemp))
                    count = count+1
                except:
                    pass
    except :
        pass

    LoudFileNameB = app.config['UPLOAD_FOLDER']+'/LB1.csv'
    try:
        with open(LoudFileNameB, newline='', errors='ignore') as lb:
            loudrows = csv.reader(lb)
            count = 10
            for loudrow in loudrows:
                try:
                    LoudTemp = loudrow[1]
                    ws.cell(row=10, column=count, value=float(LoudTemp))
                    count = count+1
                except:
                    pass
    except :
        pass
    wb.save(app.config['UPLOAD_FOLDER']+'/report.xlsx')
def temperature():
    filename = app.config['UPLOAD_FOLDER']+'/report.xlsx'
    wb = load_workbook(filename)
    ws = wb.active
    temperature_fileA = app.config['UPLOAD_FOLDER']+'/TA1.csv'
    try:
        with open(temperature_fileA, newline='', errors='ignore') as f:
            rows = csv.reader(f)
            cnt = 4
            for row in rows:
                try:
                    # 取出溫度
                    Temp1 = row[1][0:4]
                    Temp2 = row[3][0:4]
                    InsertTemp = Temp1 + ',' + Temp2
                    ws.cell(row=9, column=cnt, value=InsertTemp)
                    cnt = cnt+1
                except:
                    pass
    except:
        pass
    temperature_fileB = app.config['UPLOAD_FOLDER'] + '/TB1.csv'
    try :
        with open(temperature_fileB, newline='', errors='ignore') as f:
            rows = csv.reader(f)
            cnt = 10
            for row in rows:
                # 取出溫度
                try:
                    Temp1 = row[1][0:4]
                    Temp2 = row[3][0:4]
                    InsertTemp = Temp1 + ',' + Temp2
                    ws.cell(row=9, column=cnt, value=InsertTemp)
                    cnt = cnt + 1
                except:
                    pass
    except :
        pass
    wb.save(app.config['UPLOAD_FOLDER']+'/report.xlsx')
#電流
def electric():
    filename = app.config['UPLOAD_FOLDER']+'/report.xlsx'
    wb = load_workbook(filename)
    ws = wb.active
    EleFileNameA = app.config['UPLOAD_FOLDER']+'/EA1.csv'
    try:
        with open(EleFileNameA, newline='',errors='ignore') as e:
            reader = csv.reader((x.replace('\0', '') for x in e), delimiter='\t')
            row1 = [row for row in reader]
            for row in row1:
                try:
                    if len(row) > 1:
                        if row[0] == '1':
                            EleInsert = row[3]
                            ws.cell(row=7, column=4, value=float(EleInsert))
                        if row[0] == '2':
                            EleInsert = row[3]
                            ws.cell(row=7, column=5, value=float(EleInsert))
                        if row[0] == '3':
                            EleInsert = row[3]
                            ws.cell(row=7, column=6, value=float(EleInsert))
                        if row[0] == '4':
                            EleInsert = row[3]
                            ws.cell(row=7, column=7, value=float(EleInsert))
                        if row[0] == '5':
                            EleInsert = row[3]
                            ws.cell(row=7, column=8, value=float(EleInsert))
                except:
                    pass
    except:
        pass
    EleFileNameB = app.config['UPLOAD_FOLDER'] + '/EB1.csv'
    try:
        with open(EleFileNameB, newline='', errors='ignore') as e:
            reader = csv.reader((x.replace('\0', '') for x in e), delimiter='\t')
            row1 = [row for row in reader]
            for row in row1:
                try:
                    if len(row) > 1:
                        if row[0] == '1':
                            EleInsert = row[3]
                            ws.cell(row=7, column=10, value=float(EleInsert))
                        if row[0] == '2':
                            EleInsert = row[3]
                            ws.cell(row=7, column=11, value=float(EleInsert))
                        if row[0] == '3':
                            EleInsert = row[3]
                            ws.cell(row=7, column=12, value=float(EleInsert))
                        if row[0] == '4':
                            EleInsert = row[3]
                            ws.cell(row=7, column=13, value=float(EleInsert))
                        if row[0] == '5':
                            EleInsert = row[3]
                            ws.cell(row=7, column=14, value=float(EleInsert))
                        if row[0] == '6':
                            EleInsert = row[3]
                            ws.cell(row=7, column=15, value=float(EleInsert))
                except:
                    pass
    except:
        pass
    wb.save('/home/keit1216/keit1216_mlEngine/final_report/'+ app.config['REPORT_NAME'])

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


@app.route('/upload')
def upload_form():
    return render_template('upload.html')

@app.route('/')
def my_form():
    return render_template('key_name.html')

@app.route('/', methods=['POST'])
def my_form_post():
    text = request.form['text']
    processed_text = text.upper()
    #新增資料夾
    new_folder = ('/home/keit1216/keit1216_mlEngine/report/report_test/'+processed_text)
    os.makedirs(new_folder)
    app.config['UPLOAD_FOLDER'] = new_folder
    #將report範本複製到新的資料夾
    copyfile('/home/keit1216/keit1216_mlEngine/report/test2.xlsx', new_folder+'/test2.xlsx')
    return redirect('/upload')


@app.route('/upload', methods=['POST'])
def upload_file():
    if request.method == 'POST':
        for f in request.files.getlist('file'):
            if f and allowed_file(f.filename):
                filename = secure_filename(f.filename)
                f.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
                flash('檔案上傳成功')
        return redirect('/upload')
        # check if the post request has the file part
        '''
        if 'file' not in request.files:
            flash('No file part')
            return redirect(request.url)
        file = request.files.getlist('file')
        if file.filename == '':
            flash('No file selected for uploading')
            return redirect(request.url)
        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
            flash('檔案上傳成功')
            return redirect('/upload')
        '''

@app.route('/dropdown', methods=['GET'])
def dropdown():
    dirs = os.listdir('/home/keit1216/keit1216_mlEngine/report/report_test/')
    # This would print all the files and directories
    obj=[]
    for file in dirs:
        obj.append(file)
    return render_template('dropdown.html', obj=obj)
@app.route("/report", methods=['POST'])
def print_report():
   value = request.form.get("obj")
   #app.config['REPORT_NAME'] = 'test2.xlsx'
   app.config['REPORT_NAME'] = value + '.xlsx'
   app.config['UPLOAD_FOLDER']='/home/keit1216/keit1216_mlEngine/report/report_test/'+str(value)
   filename = app.config['UPLOAD_FOLDER'] + '/test2.xlsx'
   wb = load_workbook(filename)
   ws = wb.active
   convert()
   InsertA()
   loud()
   temperature()
   electric()
   return redirect('/download')

@app.route("/download")
def download():
    dirpath ='/home/keit1216/keit1216_mlEngine/final_report/'
    filename = app.config['REPORT_NAME']
    print(app.config['REPORT_NAME'])
    return send_file(dirpath+filename, as_attachment=True)

@app.route("/exc")
def exc():
    return redirect('/exc')

if __name__ == "__main__":
    app.run(host='0.0.0.0', port='3812')