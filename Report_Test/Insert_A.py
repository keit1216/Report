import csv
from openpyxl import load_workbook
import pandas as pd

filename = '/Users/ian/Downloads/test2.xlsx'
wb = load_workbook(filename)
ws = wb.active

#CSV to EXCEL
def convert():
    filepath_in = "/Users/ian/Downloads/AA01.csv"
    filepath_out = "/Users/ian/Downloads/excel_test.xlsx"
    pd.read_csv(filepath_in, delimiter=",").to_excel(filepath_out, index=False)
def InsertA():
    #Load excel data
    new_excel = "/Users/ian/Downloads/excel_test.xlsx"
    wb2 = load_workbook(new_excel)
    ws2 = wb2.active
    #轉速
    speed = ws2.cell(row=57, column=2).value[0:3]
    #徑向,軸向
    X_value = ws2.cell(row=59, column=2).value[0:4]
    Y_value = ws2.cell(row=59, column=3).value[0:4]
    ws.cell(row=6, column=4, value='78('+speed+')')
    ws.cell(row=11, column=4, value=X_value)
    ws.cell(row=12, column=4, value=Y_value)
#取出噪音
def loud():
    LoudFileName = '/Users/ian/Downloads/LA01test.csv'
    with open(LoudFileName, newline='', errors='ignore') as l:
        loudrows = csv.reader(l)
        count=4
        for loudrow in loudrows:
            LoudTemp = loudrow[1]
            ws.cell(row=10,column=count,value=float(LoudTemp))
            count = count+1
def temperature():
    temperature_file = '/Users/ian/Downloads/TA01.csv'
    with open(temperature_file, newline='', errors='ignore') as f:
        rows = csv.reader(f)
        cnt = 4
        for row in rows:
            # 取出溫度
            Temp1 = row[1][0:4]
            Temp2 = row[3][0:4]
            InsertTemp = Temp1 + ',' + Temp2
            ws.cell(row=9, column=cnt, value=InsertTemp)
            cnt = cnt+1
#電流
def electric():
    EleFileName = '/Users/ian/Downloads/EA01.csv'
    with open(EleFileName, newline='',errors='ignore') as e:
        reader = csv.reader((x.replace('\0', '') for x in e), delimiter='\t')
        row1 = [row for row in reader]
        for row in row1:
            if len(row) > 1:
                if row[0] == '1':
                    EleInsert = row[3]
                    ws.cell(row=7, column=4, value=float(EleInsert))
                if row[0] == '2':
                    EleInsert = row[3]
                    ws.cell(row=7, column=5, value=float(EleInsert))
                if row[0] == '3':
                    EleInsert = row[3]
                    ws.cell(row=7, column=6, value=float(EleInsert))
                if row[0] == '4':
                    EleInsert = row[3]
                    ws.cell(row=7, column=7, value=float(EleInsert))
                if row[0] == '5':
                    EleInsert = row[3]
                    ws.cell(row=7, column=8, value=float(EleInsert))

InsertA()
loud()
temperature()
electric()
wb.save('/Users/ian/Downloads/Openpyxl_test.xlsx')

