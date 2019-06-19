from openpyxl import Workbook
import csv
import pandas as pd
import re
from openpyxl import load_workbook

filename = '/Users/ian/Downloads/test.xlsx'
wb = load_workbook(filename)
ws = wb.active
#電流
def electric(filename):
    EleFileName = '/Users/ian/Downloads/EA01.csv'
    with open(EleFileName, newline='',errors='ignore') as e:
        reader = csv.reader((x.replace('\0', '') for x in e), delimiter='\t')
        row1 = [row for row in reader]
        for row in row1:
            if len(row) > 1:
                if row[0] == '1':
                    EleInsert = row[3]
                    print(EleInsert)
                    ws.cell(row=7, column=4, value=float(EleInsert))
                if row[0] == '2':
                    EleInsert = row[3]
                    print(EleInsert)
                    ws.cell(row=7, column=5, value=float(EleInsert))
                if row[0] == '3':
                    EleInsert = row[3]
                    print(EleInsert)
                    ws.cell(row=7, column=6, value=float(EleInsert))
                if row[0] == '4':
                    EleInsert = row[3]
                    print(EleInsert)
                    ws.cell(row=7, column=7, value=float(EleInsert))
                if row[0] == '5':
                    EleInsert = row[3]
                    print(EleInsert)
                    ws.cell(row=7, column=8, value=float(EleInsert))
        wb.save('/Users/ian/Downloads/Openpyxl_test.xlsx')
